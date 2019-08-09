from openpyxl import Workbook

wb = Workbook()
ws = wb.active

filename = 'blank_sample.xlsx'

ws['A1'] = 'N'
ws['B1'] = 'Name'
ws['C1'] = 'Item ID'
ws['D1'] = 'Value'
for row in range(2,400):

    ws.cell(row = row, column = 1, value = row - 1)
    ws.cell(row = row, column = 2, value = "Item" + str(row-1))
    ws.cell(row = row, column = 3, value = 999+row)
    ws.cell(row = row, column = 4, value = 520+10*row)

wb.save(filename)
