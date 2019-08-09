from openpyxl import Workbook
from openpyxl import load_workbook
from math import ceil

wb_loaded = load_workbook(filename = 'blank_sample.xlsx', read_only=True)
ws_loaded = wb_loaded.worksheets[0]
max_rows = wb_loaded.worksheets[0].max_row
num = ceil(float(max_rows)/200)

for n in range(0,int(num)):
    wb = Workbook()
    ws = wb.active

    for row in ws_loaded.iter_rows(min_row=2+n*200, max_row=1+(n+1)*200, max_col=10):
        values = (c.value for c in row)
        ws.append(values)
        
    filename = 'split_' + str(n) + '.xlsx'
    wb.save(filename)


